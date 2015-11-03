#!/usr/bin/env python
# -*- coding: utf-8 -*-

from django.views.generic import TemplateView
from mixins.mixins import AccesoMixin
from region.models import Region
from django.views.generic import CreateView
from gestor.models import Gestor, TipoGestor
from radicado.models import Radicado
from municipio.models import Municipio
from acceso.models import Actividad, Reasignados, CargaMasiva
from conf import settings
import openpyxl
from string import maketrans

from acceso.forms import ReasignacionForm, CargaMasivaForm
from .models import Evidencia
from rest_framework import mixins
from rest_framework import generics
from .serializers import EvidenciaSerializer
from django.forms.models import modelformset_factory
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponse
import time
import datetime
from openpyxl.styles import Style, PatternFill, Border, Side, Alignment, Protection, Font

from django_datatables_view.base_datatable_view import BaseDatatableView
from django.db.models import Q
from zipfile import ZipFile
from django.core.files import File
import shutil
import os

t = Style(font=Font(name='Calibri',size=12,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000'),
       fill=PatternFill(fill_type='solid',start_color='C9C9C9',end_color='FF000000'),
       alignment=Alignment(horizontal='center',vertical='center',wrap_text=True),
     number_format='General')

co = Style(font=Font(name='Calibri',size=11),
       alignment=Alignment(horizontal='center',vertical='center',wrap_text=True),
     number_format='General')

v = Style(font=Font(name='Calibri',size=12,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000'),
       fill=PatternFill(fill_type='solid',start_color='E4F5E1',end_color='FF000000'),
       alignment=Alignment(horizontal='center',vertical='center',wrap_text=True),
     number_format='General')

vc = Style(font=Font(name='Calibri',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000'),
       fill=PatternFill(fill_type='solid',start_color='D4F5CE',end_color='FF000000'),
       alignment=Alignment(horizontal='center',vertical='center',wrap_text=True),
     number_format='General')

def encode_cp437(s, _noqmarks=maketrans('?', '_')):
    return s.encode('cp437', errors='replace').translate(_noqmarks)

class EvidenciaViewSet(mixins.UpdateModelMixin,generics.GenericAPIView):
    queryset = Evidencia.objects.all()
    serializer_class = EvidenciaSerializer
    lookup_url_kwarg = 'id_evidencia'


    def put(self, request, *args, **kwargs):
        return self.update(request, *args, **kwargs)

class AccesoView(AccesoMixin,TemplateView):
    template_name = 'acceso_tipo.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        return super(AccesoView,self).get_context_data(**kwargs)

class AccesoTipoView(AccesoMixin,TemplateView):
    template_name = 'acceso.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['TIPO'] = TipoGestor.objects.get(pk=self.kwargs['tipo_gestor']).tipo
        return super(AccesoTipoView,self).get_context_data(**kwargs)

class AccesoCalificacionView(AccesoMixin,TemplateView):
    template_name = 'acceso_gestores.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['ID_TIPO_GESTOR'] = self.kwargs['tipo_gestor']
        kwargs['TIPO'] = TipoGestor.objects.get(pk=self.kwargs['tipo_gestor']).tipo
        return super(AccesoCalificacionView,self).get_context_data(**kwargs)

class AccesoCalificacionTotalView(AccesoMixin,TemplateView):
    template_name = 'acceso_radicados_total.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['ID_TIPO_GESTOR'] = self.kwargs['tipo_gestor']
        kwargs['TIPO'] = TipoGestor.objects.get(pk=self.kwargs['tipo_gestor']).tipo
        return super(AccesoCalificacionTotalView,self).get_context_data(**kwargs)


class AccesoListadoRadicadosView(AccesoMixin,TemplateView):
    template_name = 'acceso_radicados.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['NOMBRE'] = Gestor.objects.get(pk=self.kwargs['id_gestor']).nombre
        kwargs['MUNICIPIO'] = Municipio.objects.get(pk=self.kwargs['id_municipio']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['ID_GESTOR'] = self.kwargs['id_gestor']
        kwargs['ID_MUNICIPIO'] = self.kwargs['id_municipio']
        kwargs['ID_TIPO_GESTOR'] = self.kwargs['tipo_gestor']
        kwargs['TIPO'] = TipoGestor.objects.get(pk=self.kwargs['tipo_gestor']).tipo
        return super(AccesoListadoRadicadosView,self).get_context_data(**kwargs)

class AccesoListadoMunicipiosView(AccesoMixin,TemplateView):
    template_name = 'acceso_municipios.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['NOMBRE'] = Gestor.objects.get(pk=self.kwargs['id_gestor']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['ID_GESTOR'] = self.kwargs['id_gestor']
        kwargs['ID_TIPO_GESTOR'] = self.kwargs['tipo_gestor']
        kwargs['TIPO'] = TipoGestor.objects.get(pk=self.kwargs['tipo_gestor']).tipo
        return super(AccesoListadoMunicipiosView,self).get_context_data(**kwargs)

def evidencia_form(request,id_radicado,pk,id_gestor,id_municipio,tipo_gestor):
    EvidenciaFormSet = modelformset_factory(Evidencia, fields=('soporte',),extra=0)
    if request.method == "POST":
        formset = EvidenciaFormSet(request.POST, request.FILES, queryset=Evidencia.objects.filter(radicado__id=id_radicado).filter(gestor__id=id_gestor))
        if formset.is_valid():
            formset.save()
            for form in formset.forms:
                if form.cleaned_data['soporte'] != None:
                    obj = Evidencia.objects.get(pk=form.instance.id)
                    obj.usuario = request.user
                    obj.modificacion = datetime.datetime.now()
                    obj.save()
    else:
        formset = EvidenciaFormSet(queryset=Evidencia.objects.filter(radicado__id=id_radicado).filter(gestor__id=id_gestor),)
    return render_to_response("evidencias_radicado.html",{"formset":formset,"user":request.user,"REGION":Region.objects.get(pk=pk).nombre,
                                                          "gestor":Gestor.objects.get(pk=id_gestor).nombre,
                                                          "radicado":Radicado.objects.get(pk=id_radicado),
                                                          "municipio":Municipio.objects.get(pk=id_municipio),
                                                          "TIPO":TipoGestor.objects.get(pk=tipo_gestor).tipo},
                              context_instance=RequestContext(request))

def evidencia_total_form(request,id_radicado,pk,tipo_gestor):
    EvidenciaFormSet = modelformset_factory(Evidencia, fields=('soporte',),extra=0)
    if request.method == "POST":
        formset = EvidenciaFormSet(request.POST, request.FILES, queryset=Evidencia.objects.filter(radicado__id=id_radicado))
        if formset.is_valid():
            formset.save()
            for form in formset.forms:
                if form.cleaned_data['soporte'] != None:
                    obj = Evidencia.objects.get(pk=form.instance.id)
                    obj.usuario = request.user
                    obj.modificacion = datetime.datetime.now()
                    obj.save()
    else:
        formset = EvidenciaFormSet(queryset=Evidencia.objects.filter(radicado__id=id_radicado),)
    return render_to_response("evidencias_radicado_total.html",{"formset":formset,"user":request.user,
                                                                "REGION":Region.objects.get(pk=pk).nombre,
                                                                "radicado":Radicado.objects.get(pk=id_radicado),
                                                                "TIPO":TipoGestor.objects.get(pk=tipo_gestor).tipo},context_instance=RequestContext(request))

def reporte_acceso(request,pk,id_gestor,tipo_gestor):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Actividades Gestores.xlsx'
    archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

    logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
    logo.drawing.top = 10
    logo.drawing.left = 25

    hoja1 = archivo.get_sheet_by_name('hoja1')
    hoja1.title = "Informe Gestores"
    hoja1.add_image(logo)

    celda = hoja1.cell('E2')
    celda.value = 'ACCESO'

    celda = hoja1.cell('E3')
    celda.value = 'REPORTE GESTORES'

    celda = hoja1.cell('I3')
    celda.value = time.strftime("%d/%m/%y")

    celda = hoja1.cell('I4')
    celda.value = time.strftime("%I:%M:%S %p")

    row_num = 5

    columnas = Actividad.objects.order_by('id').values('nombre','titulo')
    columns = [tuple(['Radicado',30]),
               tuple(['Departamento',30]),
               tuple(['Municipio',30]),
               tuple(['Gestor',30]),
               tuple(['Ciclo',30]),
               tuple(['Componente',30]),
               tuple(['Modulo',30]),
               tuple(['Actividad',30]),
               tuple(['ID',30]),
               tuple(['Encargado',30]),
               tuple(['Valor',30]),
               tuple(['Soporte',30]),
               tuple(['Corte',30]),
               ]



    for col_num in xrange(len(columns)):
        c = hoja1.cell(row=row_num, column=col_num+1)
        c.value = columns[col_num][0]
        c.style = t
        hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]


    evidencias = Evidencia.objects.filter(radicado__region__id=pk).filter(gestor__id=id_gestor)

    for evidencia in evidencias:
        row_num += 1
        row = [
            evidencia.radicado.numero,
            evidencia.radicado.municipio.departamento.nombre,
            evidencia.radicado.municipio.nombre,
            evidencia.gestor.nombre,
            evidencia.ciclo.nombre,
            evidencia.componente.nombre,
            evidencia.modulo.nombre,
            str(evidencia.actividad.nombre)+" - "+str(evidencia.actividad.titulo),
            evidencia.actividad.id,
            evidencia.encargado.encargado,
            evidencia.valor.valor,
            str(evidencia.soporte),
            str(evidencia.corte.fecha)+" - "+str(evidencia.corte.titulo) if evidencia.corte != None else "",
        ]

        for col_num in xrange(len(row)):
            c = hoja1.cell(row=row_num, column=col_num+1)
            if row[col_num] == True:
                c.value = "SI"
            if row[col_num] == False:
                c.value = "NO"
            if row[col_num] == None:
                c.value = ""
            else:
                c.value = row[col_num]
            c.style = co



    archivo.save(response)
    return response

def reporte_total(request,pk,tipo_gestor):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Actividades Gestores.xlsx'
    archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

    logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
    logo.drawing.top = 10
    logo.drawing.left = 25

    hoja1 = archivo.get_sheet_by_name('hoja1')
    hoja1.title = "Informe Gestores"
    hoja1.add_image(logo)

    celda = hoja1.cell('E2')
    celda.value = 'ACCESO'

    celda = hoja1.cell('E3')
    celda.value = 'REPORTE GESTORES'

    celda = hoja1.cell('I3')
    celda.value = time.strftime("%d/%m/%y")

    celda = hoja1.cell('I4')
    celda.value = time.strftime("%I:%M:%S %p")

    row_num = 5

    columnas = Actividad.objects.order_by('id').values('nombre','titulo')
    columns = [tuple(['Radicado',30]),
               tuple(['Departamento',30]),
               tuple(['Municipio',30]),
               tuple(['Gestor',30]),
               tuple(['Ciclo',30]),
               tuple(['Componente',30]),
               tuple(['Modulo',30]),
               tuple(['Actividad',30]),
               tuple(['ID',30]),
               tuple(['Encargado',30]),
               tuple(['Valor',30]),
               tuple(['Soporte',30]),
               tuple(['Corte',30]),
               ]



    for col_num in xrange(len(columns)):
        c = hoja1.cell(row=row_num, column=col_num+1)
        c.value = columns[col_num][0]
        c.style = t
        hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]


    evidencias = Evidencia.objects.filter(radicado__region__id=pk)

    for evidencia in evidencias:
        row_num += 1
        row = [
            evidencia.radicado.numero,
            evidencia.radicado.municipio.departamento.nombre,
            evidencia.radicado.municipio.nombre,
            evidencia.gestor.nombre,
            evidencia.ciclo.nombre,
            evidencia.componente.nombre,
            evidencia.modulo.nombre,
            str(evidencia.actividad.nombre)+" - "+str(evidencia.actividad.titulo),
            evidencia.actividad.id,
            evidencia.encargado.encargado,
            evidencia.valor.valor,
            str(evidencia.soporte),
            str(evidencia.corte.fecha)+" - "+str(evidencia.corte.titulo) if evidencia.corte != None else "",
        ]

        for col_num in xrange(len(row)):
            c = hoja1.cell(row=row_num, column=col_num+1)
            if row[col_num] == True:
                c.value = "SI"
            if row[col_num] == False:
                c.value = "NO"
            if row[col_num] == None:
                c.value = ""
            else:
                c.value = row[col_num]
            c.style = co



    archivo.save(response)
    return response

class ReasignarView(AccesoMixin,CreateView):
    model = Reasignados
    form_class = ReasignacionForm
    template_name = "formulario_reasignados.html"
    success_url = "../../../"

class MasivoView(AccesoMixin,TemplateView):
    template_name = 'acceso_masivo.html'

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['ID_TIPO_GESTOR'] = self.kwargs['tipo_gestor']
        kwargs['TIPO'] = TipoGestor.objects.get(pk=self.kwargs['tipo_gestor']).tipo
        return super(MasivoView,self).get_context_data(**kwargs)

class MasivoTableView(BaseDatatableView):
    model = CargaMasiva
    columns = [
        'id',
        'region',
        'fecha',
        'usuario',
        'excel',
        'archivo',
    ]

    order_columns = [
        'id',
        'region',
        'fecha',
        'usuario',
        'excel',
        'archivo',
    ]

    def get_initial_queryset(self):
        if not self.model:
            raise NotImplementedError("Need to provide a model or implement get_initial_queryset!")
        return self.model.objects.all().filter(region__id=self.kwargs['pk'])

    def filter_queryset(self, qs):
        search = self.request.GET.get(u'search[value]', None)
        q = Q()
        if search:
            q |= Q(**{'usuario__icontains' : search})
            qs = qs.filter(q)
        return qs

    def render_column(self, row, column):
        if column == 'region':
            return row.region.nombre
        if column == 'fecha':
            return row.fecha.strftime('%d/%m/%Y --- %I:%M:%S %p')
        if column == 'usuario':
            return row.usuario.username
        if column == 'excel':
            return str(row.excel)
        if column == 'archivo':
            return str(row.archivo)
        else:
            return super(MasivoTableView,self).render_column(row,column)

class MasivoNuevoView(AccesoMixin,CreateView):
    model = CargaMasiva
    form_class = CargaMasivaForm
    template_name = "formulario_carga_masiva.html"
    success_url = "../"

    def get_context_data(self, **kwargs):
        kwargs['REGION'] = Region.objects.get(pk=self.kwargs['pk']).nombre
        kwargs['ID_REGION'] = self.kwargs['pk']
        kwargs['ID_TIPO_GESTOR'] = self.kwargs['tipo_gestor']
        kwargs['TIPO'] = TipoGestor.objects.get(pk=self.kwargs['tipo_gestor']).tipo
        return super(MasivoNuevoView,self).get_context_data(**kwargs)

def ejecutar_masivo(request,pk,id_masivo,tipo_gestor):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Carga Masiva.xlsx'
    archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

    logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
    logo.drawing.top = 10
    logo.drawing.left = 25

    hoja1 = archivo.get_sheet_by_name('hoja1')
    hoja1.title = "Reporte Carga Masiva"
    hoja1.add_image(logo)

    celda = hoja1.cell('E2')
    celda.value = 'ACCESO'

    celda = hoja1.cell('E3')
    celda.value = 'Reporte Carga Masiva'

    celda = hoja1.cell('I3')
    celda.value = time.strftime("%d/%m/%y")

    celda = hoja1.cell('I4')
    celda.value = time.strftime("%I:%M:%S %p")

    row_num = 5

    columns = [tuple(['RADICADO',30]),
               tuple(['ID ACTIVIDAD',30]),
               tuple(['PATH RELATIVO',30]),
               tuple(['CARGADO',30]),
               tuple(['INFORMACION',60]),
               tuple(['GESTOR',60]),
               tuple(['CEDULA',60]),
               ]

    for col_num in xrange(len(columns)):
        c = hoja1.cell(row=row_num, column=col_num+1)
        c.value = columns[col_num][0]
        c.style = t
        hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]


    masivo = CargaMasiva.objects.get(pk=id_masivo)
    x=settings.MEDIA_ROOT+'/'+str(masivo.archivo)
    soportes = ZipFile(settings.MEDIA_ROOT+'//'+str(masivo.archivo),'r')

    archivo_masivo = openpyxl.load_workbook(settings.MEDIA_ROOT+'/'+str(masivo.excel))

    hoja1_masivo = archivo_masivo.get_sheet_by_name('Hoja1')

    i = 0

    for fila in hoja1_masivo.rows:
        i += 1
        if i > 2:
            proceso =""
            gestor = ""
            cedula = ""

            evidencia = Evidencia.objects.filter(radicado__numero=fila[0].value).filter(actividad__id=fila[1].value)


            if len(evidencia) == 0:
                proceso = "No existe el Radicado"
            if len(evidencia) == 1:
                #if evidencia[0].soporte == "":
                #    e = evidencia[0]
                #    e.soporte = File(open("C://Temp//pendiente.txt", 'rb'))
                #    e.save()
                #    proceso = "Cargado con exito"
                #else:
                #    proceso = "Archivo Cargado Anteriormente"
                gestor = evidencia[0].gestor.nombre
                cedula = evidencia[0].gestor.cedula
                try:
                    info = soportes.getinfo(fila[2].value)
                except:
                    try:
                        info = soportes.getinfo(encode_cp437(fila[2].value))
                    except:
                        proceso = "No Existe el archivo en el path"
                    else:
                        proceso = "Soporte cargado"
                        soportes.extract(encode_cp437(fila[2].value),"C:\Temp")
                        e = evidencia[0]
                        e.soporte = File(open("C://Temp//" + encode_cp437(fila[2].value), 'rb'))
                        e.save()
                        os.remove("C://Temp//" + encode_cp437(fila[2].value))
                else:
                    proceso = "Soporte cargado"
                    soportes.extract(fila[2].value,"C:\Temp")
                    e = evidencia[0]
                    e.soporte = File(open("C://Temp//" + fila[2].value, 'rb'))
                    e.save()
                    os.remove("C://Temp//" + fila[2].value)
            if len(evidencia) >= 2:
                proceso = "Se encontro mas de un radicado con el mismo numero"


            row_num += 1
            row = [
                fila[0].value,
                fila[1].value,
                fila[2].value,
                proceso,
                gestor,
                cedula,
            ]

            for col_num in xrange(len(row)):
                c = hoja1.cell(row=row_num, column=col_num+1)
                if row[col_num] == True:
                    c.value = "SI"
                if row[col_num] == False:
                    c.value = "NO"
                if row[col_num] == None:
                    c.value = ""
                else:
                    c.value = row[col_num]
                c.style = co

    archivo.save(response)
    return response