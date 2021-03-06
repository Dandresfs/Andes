from django.contrib import admin
from .models import Masivo, Actividad, Entregable, Grupo, SoporteEntregableEscuelaTic, ParticipanteEscuelaTic, EvidenciaEscuelaTic, Valor, Corte
from .models import RadicadoFormacion, ParticipanteDocente, ActividadDocentes, EntregableDocentes, GrupoDocentes, SoporteEntregableDocente, EvidenciaDocentes, ValorDocente, CorteDocente
from .models import CargasMasivas, AreaCurricular,Grado, Genero, Competencias, GrupoPoblacional
from formador.models import Formador
from municipio.models import Municipio
from django.http import HttpResponse
from conf import settings
import openpyxl
import time
import datetime
from openpyxl.styles import Style, PatternFill, Border, Side, Alignment, Protection, Font
from zipfile import ZipFile
from truchas.models import ParticipanteEscuelaTicTrucho, CodigoMasivo
from formacion.models import SoporteEntregableEscuelaTic, ParticipanteEscuelaTic, EvidenciaDocentes
import os
import shutil
from django.core.files import File
from .models import RevisionInterventoriaEscuelaTic, RevisionInterventoriaDocente, RevisionInterventoriaDocenteSoporte, RevisionInterventoriaEscuelaTicSoporte
from .models import RevisionInterventoriaDocenteSoporteActividades, RevisionInterventoriaEscuelaTicSoporteActividades
from django.shortcuts import HttpResponseRedirect
from formacion.models import AreaCurricular,Grado,Genero,Competencias, GrupoPoblacional, CargaMasivaListados

t = Style(font=Font(name='Calibri',size=12,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000'),
       fill=PatternFill(fill_type='solid',start_color='C9C9C9',end_color='FF000000'),
       alignment=Alignment(horizontal='center',vertical='center',wrap_text=True),
     number_format='General')

co = Style(font=Font(name='Calibri',size=11),
       alignment=Alignment(horizontal='center',vertical='center',wrap_text=True),
     number_format='General')

v = Style(font=Font(name='Calibri',size=12,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000'),
       fill=PatternFill(fill_type='solid',start_color='E4F5E1',end_color='FF000000'),
       alignment=Alignment(horizontal='center',vertical='center',wrap_text=True),
     number_format='General')

vc = Style(font=Font(name='Calibri',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='FF000000'),
       fill=PatternFill(fill_type='solid',start_color='D4F5CE',end_color='FF000000'),
       alignment=Alignment(horizontal='center',vertical='center',wrap_text=True),
     number_format='General')

admin.site.register(Masivo)
admin.site.register(Actividad)
admin.site.register(ActividadDocentes)
admin.site.register(Grupo)
admin.site.register(GrupoDocentes)
admin.site.register(SoporteEntregableEscuelaTic)
admin.site.register(SoporteEntregableDocente)
admin.site.register(ParticipanteEscuelaTic)
admin.site.register(ParticipanteDocente)
admin.site.register(EvidenciaEscuelaTic)
admin.site.register(EvidenciaDocentes)
admin.site.register(Valor)
admin.site.register(ValorDocente)
admin.site.register(Corte)
admin.site.register(CorteDocente)
admin.site.register(RadicadoFormacion)

def validateEmail( email ):
    from django.core.validators import validate_email
    from django.core.exceptions import ValidationError
    try:
        validate_email( email )
        return True
    except ValidationError:
        return False

def carga_grupos(modeladmin,request,queryset):
    for archivo_queryset in queryset:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Carga Masiva.xlsx'
        archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

        logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
        logo.drawing.top = 10
        logo.drawing.left = 25

        hoja1 = archivo.get_sheet_by_name('hoja1')
        hoja1.title = "Reporte Carga Masiva Grupos"
        hoja1.add_image(logo)

        celda = hoja1.cell('E2')
        celda.value = 'Formacion'

        celda = hoja1.cell('E3')
        celda.value = 'Reporte Carga Masiva Grupos'

        celda = hoja1.cell('I3')
        celda.value = time.strftime("%d/%m/%y")

        celda = hoja1.cell('I4')
        celda.value = time.strftime("%I:%M:%S %p")

        row_num = 5

        columns = [tuple(['REGION',30]),
                   tuple(['DEPARTAMENTO',30]),
                   tuple(['MUNICIPIO',30]),
                   tuple(['LUGAR DE REUNION',60]),
                   tuple(['CODIGO GRUPO',30]),
                   tuple(['NOMBRE FORMADOR',60]),
                   tuple(['CEDULA',30]),
                   tuple(['RESULTADO CARGA',30]),
                   ]

        for col_num in xrange(len(columns)):
            c = hoja1.cell(row=row_num, column=col_num+1)
            c.value = columns[col_num][0]
            c.style = t
            hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]


        archivo_masivo = openpyxl.load_workbook(settings.MEDIA_ROOT+'/'+unicode(archivo_queryset.archivo))

        hoja1_masivo = archivo_masivo.get_sheet_by_name('Hoja1')

        i = 0

        for fila in hoja1_masivo.rows:
            i += 1
            if i > 1:
                proceso =""

                if fila[6].value != None:
                    try:
                        cedula = long(fila[6].value)
                    except ValueError:
                        proceso = "El numero de cedula solo debe contener numeros"
                    else:
                        if Formador.objects.filter(cedula=cedula).count() == 1:
                            formador = Formador.objects.get(cedula=cedula)
                            if Municipio.objects.filter(nombre=fila[2].value).count() == 1:
                                municipio = Municipio.objects.get(nombre=fila[2].value)
                                if fila[4].value != None:
                                    if Grupo.objects.filter(nombre=fila[4].value).count() == 0:
                                        grupo = Grupo(formador=formador,nombre=fila[4].value,municipio=municipio,direccion=fila[3].value)
                                        grupo.save()
                                        proceso = "Grupo creado correctamente"
                                    else:
                                        proceso = "El grupo ya existe"
                                else:
                                    proceso = "Esta vacio el codigo de grupo"
                            else:
                                proceso = "No existe el municipio"
                        else:
                            proceso = "No hay un solo formador registrado con el numero de cedula"
                else:
                    proceso = "No hay numero de cedula valido"

                row_num += 1
                row = [
                    fila[0].value,
                    fila[1].value,
                    fila[2].value,
                    fila[3].value,
                    fila[4].value,
                    fila[5].value,
                    fila[6].value,
                    proceso
                ]

                for col_num in xrange(len(row)):
                    c = hoja1.cell(row=row_num, column=col_num+1)
                    if row[col_num] == True:
                        c.value = "SI"
                    if row[col_num] == False:
                        c.value = "NO"
                    if row[col_num] == None:
                        c.value = ""
                    else:
                        c.value = row[col_num]
                    c.style = co

        archivo.save(response)
        return response
carga_grupos.short_description = "Cargar grupos"

def carga_participantes(modeladmin,request,queryset):
    for archivo_queryset in queryset:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Carga Masiva.xlsx'
        archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

        logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
        logo.drawing.top = 10
        logo.drawing.left = 25

        hoja1 = archivo.get_sheet_by_name('hoja1')
        hoja1.title = "Carga Masiva Participantes"
        hoja1.add_image(logo)

        celda = hoja1.cell('E2')
        celda.value = 'Formacion'

        celda = hoja1.cell('E3')
        celda.value = 'Carga Masiva Participantes'

        celda = hoja1.cell('I3')
        celda.value = time.strftime("%d/%m/%y")

        celda = hoja1.cell('I4')
        celda.value = time.strftime("%I:%M:%S %p")

        row_num = 5

        columns = [tuple(['CODIGO GRUPO',30]),
                   tuple(['NOMBRES',30]),
                   tuple(['APELLIDOS',30]),
                   tuple(['CEDULA',60]),
                   tuple(['GENERO',30]),
                   tuple(['NIVEL EDUCATIVO',60]),
                   tuple(['CELULAR',30]),
                   tuple(['CORREO',30]),
                   tuple(['TIPO POBLACION',30]),
                   tuple(['CODIGO ANSPE',30]),
                   tuple(['TIPO PROYECTO',30]),
                   tuple(['GRUPO DE CONFORMACION',30]),
                   ]

        for col_num in xrange(len(columns)):
            c = hoja1.cell(row=row_num, column=col_num+1)
            c.value = columns[col_num][0]
            c.style = t
            hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]


        archivo_masivo = openpyxl.load_workbook(settings.MEDIA_ROOT+'/'+unicode(archivo_queryset.archivo))

        hoja1_masivo = archivo_masivo.get_sheet_by_name('Hoja1')

        i = 0

        for fila in hoja1_masivo.rows:
            i += 1
            if i > 1:
                proceso =""
                try:
                    cedula = fila[3].value.replace('.','').replace(',','')
                except:
                    cedula = fila[3].value
                nombres = fila[1].value
                apellidos = fila[2].value


                if cedula == "" or cedula == None:
                    proceso = "El campo de cedula esta vacio"
                else:
                    try:
                        long(cedula)
                    except ValueError:
                        proceso = "El numero de cedula es invalido"
                    else:
                        if ParticipanteEscuelaTic.objects.filter(cedula=cedula).count() != 0:
                            proceso = "La Cedula ya se encuentra registrada"
                        else:
                            if nombres == "" or nombres == None:
                                proceso = "El campo de nombres esta vacio"
                            else:
                                if apellidos == "" or apellidos == None:
                                    proceso = "El campo de apellidos esta vacio"
                                else:
                                    if fila[0].value == "" or fila[3].value == None:
                                        proceso = "El Codigo de grupo no es valido"
                                    else:
                                        if Grupo.objects.filter(nombre=fila[0].value).count() != 1:
                                            proceso = "No hay un solo grupo registrado con el codigo de grupo"
                                        else:
                                            grupo = Grupo.objects.get(nombre=fila[0].value)
                                            proceso = "Registrado Correctamente"
                                            nuevo = ParticipanteEscuelaTic()
                                            nuevo.grupo = grupo
                                            nuevo.formador = grupo.formador
                                            nuevo.nombres = fila[1].value
                                            nuevo.apellidos = fila[2].value
                                            nuevo.cedula = long(cedula)
                                            if fila[4].value == None or fila[4].value == "":
                                                nuevo.genero = "Masculino"
                                                proceso = "Registrado Correctamente - Genero Masculino por defecto"
                                            else:
                                                nuevo.genero = fila[4].value
                                            nuevo.nivel_educativo = fila[5].value
                                            nuevo.telefono = fila[6].value
                                            if validateEmail(fila[7].value):
                                                nuevo.correo = fila[7].value
                                            nuevo.poblacion = fila[8].value
                                            nuevo.codigo_anspe = fila[9].value
                                            nuevo.tipo_proyecto = fila[10].value
                                            nuevo.grupo_conformacion = fila[11].value
                                            nuevo.save()

                row_num += 1
                row = [
                    fila[0].value,
                    fila[1].value,
                    fila[2].value,
                    fila[3].value,
                    fila[4].value,
                    fila[5].value,
                    fila[6].value,
                    fila[7].value,
                    fila[8].value,
                    fila[9].value,
                    fila[10].value,
                    fila[11].value,
                    proceso
                ]

                for col_num in xrange(len(row)):
                    c = hoja1.cell(row=row_num, column=col_num+1)
                    if row[col_num] == True:
                        c.value = "SI"
                    if row[col_num] == False:
                        c.value = "NO"
                    if row[col_num] == None:
                        c.value = ""
                    else:
                        c.value = row[col_num]
                    c.style = co

        archivo.save(response)
        return response
carga_participantes.short_description = "Cargar participantes"

def carga_radicados(modeladmin,request,queryset):
    for archivo_queryset in queryset:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Carga Masiva.xlsx'
        archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

        logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
        logo.drawing.top = 10
        logo.drawing.left = 25

        hoja1 = archivo.get_sheet_by_name('hoja1')
        hoja1.title = "Carga Masiva Radicados"
        hoja1.add_image(logo)

        celda = hoja1.cell('E2')
        celda.value = 'Formacion'

        celda = hoja1.cell('E3')
        celda.value = 'Carga Masiva Radicados'

        celda = hoja1.cell('I3')
        celda.value = time.strftime("%d/%m/%y")

        celda = hoja1.cell('I4')
        celda.value = time.strftime("%I:%M:%S %p")

        row_num = 5

        columns = [tuple(['RADICADO',30]),
                   tuple(['CODIGO DANE INSTITUCION',30]),
                   tuple(['NOMBRE INSTITUCION',30]),
                   tuple(['CODIGO DANE SEDE EDUCATIVA',30]),
                   tuple(['NOMBRE DE LA SEDE EDUCATIVA',30]),
                   ]

        for col_num in xrange(len(columns)):
            c = hoja1.cell(row=row_num, column=col_num+1)
            c.value = columns[col_num][0]
            c.style = t
            hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]


        archivo_masivo = openpyxl.load_workbook(settings.MEDIA_ROOT+'/'+unicode(archivo_queryset.archivo))

        hoja1_masivo = archivo_masivo.get_sheet_by_name('Hoja1')

        i = 0

        for fila in hoja1_masivo.rows:
            i += 1
            if i > 1:
                proceso =""
                if fila[0].value != None:
                    try:
                        radicado = int(fila[0].value)
                    except ValueError:
                        proceso = "El numero de radicado solo debe contener numeros"
                    else:
                        if RadicadoFormacion.objects.filter(numero=radicado).count() == 0:
                            nuevo = RadicadoFormacion()
                            nuevo.numero = radicado
                            if fila[1].value != None:
                                nuevo.dane_ie = fila[1].value
                            if fila[2].value != None:
                                nuevo.nombre_ie = fila[2].value
                            if fila[3].value != None:
                                nuevo.dane_sede = fila[3].value
                            if fila[4].value != None:
                                nuevo.nombre_sede = fila[4].value
                            nuevo.save()
                            proceso = "Radicado creado satisfactoriamente"
                        else:
                            proceso = "Ya existe el radicado"
                else:
                    proceso = "No hay numero de radicado valido"

                row_num += 1
                row = [
                    fila[0].value,
                    fila[1].value,
                    fila[2].value,
                    fila[3].value,
                    fila[4].value,
                    proceso
                ]

                for col_num in xrange(len(row)):
                    c = hoja1.cell(row=row_num, column=col_num+1)
                    if row[col_num] == True:
                        c.value = "SI"
                    if row[col_num] == False:
                        c.value = "NO"
                    if row[col_num] == None:
                        c.value = ""
                    else:
                        c.value = row[col_num]
                    c.style = co

        archivo.save(response)
        return response
carga_radicados.short_description = "Cargar radicados"

def actualizar_radicados(modeladmin,request,queryset):
    for archivo_queryset in queryset:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Actualizacion Radicados.xlsx'
        archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

        logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
        logo.drawing.top = 10
        logo.drawing.left = 25

        hoja1 = archivo.get_sheet_by_name('hoja1')
        hoja1.title = "Actualizacion Radicados"
        hoja1.add_image(logo)

        celda = hoja1.cell('E2')
        celda.value = 'Formacion'

        celda = hoja1.cell('E3')
        celda.value = 'Actualizacion Radicados'

        celda = hoja1.cell('I3')
        celda.value = time.strftime("%d/%m/%y")

        celda = hoja1.cell('I4')
        celda.value = time.strftime("%I:%M:%S %p")

        row_num = 5

        columns = [tuple(['RADICADO',30]),
                   tuple(['CODIGO DANE INSTITUCION',30]),
                   tuple(['NOMBRE INSTITUCION',30]),
                   tuple(['CODIGO DANE SEDE EDUCATIVA',30]),
                   tuple(['NOMBRE DE LA SEDE EDUCATIVA',30]),
                   ]

        for col_num in xrange(len(columns)):
            c = hoja1.cell(row=row_num, column=col_num+1)
            c.value = columns[col_num][0]
            c.style = t
            hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]


        archivo_masivo = openpyxl.load_workbook(settings.MEDIA_ROOT+'/'+unicode(archivo_queryset.archivo))

        hoja1_masivo = archivo_masivo.get_sheet_by_name('Hoja1')

        i = 0

        for fila in hoja1_masivo.rows:
            i += 1
            if i > 1:
                proceso =""
                if fila[0].value != None:
                    try:
                        radicado = int(fila[0].value)
                    except ValueError:
                        proceso = "El numero de radicado solo debe contener numeros"
                    else:
                        if RadicadoFormacion.objects.filter(numero=radicado).count() == 1:
                            radicado = RadicadoFormacion.objects.get(numero=radicado)
                            if fila[1].value != None:
                                radicado.dane_ie = fila[1].value
                            if fila[2].value != None:
                                radicado.nombre_ie = fila[2].value
                            if fila[3].value != None:
                                radicado.dane_sede = fila[3].value
                            if fila[4].value != None:
                                radicado.nombre_sede = fila[4].value
                            radicado.save()
                            proceso = "Radicado actualizado satisfactoriamente"
                        else:
                            proceso = "Hay mas de un radicado con este numero"
                else:
                    proceso = "No hay numero de radicado valido"

                row_num += 1
                row = [
                    fila[0].value,
                    fila[1].value,
                    fila[2].value,
                    fila[3].value,
                    fila[4].value,
                    proceso
                ]

                for col_num in xrange(len(row)):
                    c = hoja1.cell(row=row_num, column=col_num+1)
                    if row[col_num] == True:
                        c.value = "SI"
                    if row[col_num] == False:
                        c.value = "NO"
                    if row[col_num] == None:
                        c.value = ""
                    else:
                        c.value = row[col_num]
                    c.style = co

        archivo.save(response)
        return response
actualizar_radicados.short_description = "Actualizar radicados"

def carga_grupos_docentes(modeladmin,request,queryset):
    for archivo_queryset in queryset:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Carga Masiva.xlsx'
        archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

        logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
        logo.drawing.top = 10
        logo.drawing.left = 25

        hoja1 = archivo.get_sheet_by_name('hoja1')
        hoja1.title = "Reporte Carga Masiva Grupos"
        hoja1.add_image(logo)

        celda = hoja1.cell('E2')
        celda.value = 'Formacion'

        celda = hoja1.cell('E3')
        celda.value = 'Reporte Carga Masiva Grupos'

        celda = hoja1.cell('I3')
        celda.value = time.strftime("%d/%m/%y")

        celda = hoja1.cell('I4')
        celda.value = time.strftime("%I:%M:%S %p")

        row_num = 5

        columns = [tuple(['REGION',30]),
                   tuple(['DEPARTAMENTO',30]),
                   tuple(['SECRETARIA',30]),
                   tuple(['LUGAR DE REUNION',60]),
                   tuple(['CODIGO GRUPO',30]),
                   tuple(['NOMBRE FORMADOR',60]),
                   tuple(['CEDULA',30]),
                   tuple(['RESULTADO CARGA',30]),
                   ]

        for col_num in xrange(len(columns)):
            c = hoja1.cell(row=row_num, column=col_num+1)
            c.value = columns[col_num][0]
            c.style = t
            hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]


        archivo_masivo = openpyxl.load_workbook(settings.MEDIA_ROOT+'/'+unicode(archivo_queryset.archivo))

        hoja1_masivo = archivo_masivo.get_sheet_by_name('Hoja1')

        i = 0

        for fila in hoja1_masivo.rows:
            i += 1
            if i > 1:
                proceso =""

                if fila[6].value != None:
                    try:
                        cedula = long(fila[6].value)
                    except ValueError:
                        proceso = "El numero de cedula solo debe contener numeros"
                    else:
                        if Formador.objects.filter(cedula=cedula).count() == 1:
                            formador = Formador.objects.get(cedula=cedula)
                            if Municipio.objects.filter(nombre=fila[2].value).count() == 1:
                                municipio = Municipio.objects.get(nombre=fila[2].value)
                                if fila[4].value != None:
                                    if GrupoDocentes.objects.filter(nombre=fila[4].value).count() == 0:
                                        grupo = GrupoDocentes(formador=formador,nombre=fila[4].value,municipio=municipio,direccion=fila[3].value)
                                        grupo.save()
                                        proceso = "Grupo creado correctamente"
                                    else:
                                        proceso = "El grupo ya existe"
                                else:
                                    proceso = "Esta vacio el codigo de grupo"
                            else:
                                proceso = "No existe la secretaria"
                        else:
                            proceso = "No hay un solo formador registrado con el numero de cedula"
                else:
                    proceso = "No hay numero de cedula valido"

                row_num += 1
                row = [
                    fila[0].value,
                    fila[1].value,
                    fila[2].value,
                    fila[3].value,
                    fila[4].value,
                    fila[5].value,
                    fila[6].value,
                    proceso
                ]

                for col_num in xrange(len(row)):
                    c = hoja1.cell(row=row_num, column=col_num+1)
                    if row[col_num] == True:
                        c.value = "SI"
                    if row[col_num] == False:
                        c.value = "NO"
                    if row[col_num] == None:
                        c.value = ""
                    else:
                        c.value = row[col_num]
                    c.style = co

        archivo.save(response)
        return response
carga_grupos_docentes.short_description = "Cargar grupos de Docentes"

def carga_docentes(modeladmin,request,queryset):
    for archivo_queryset in queryset:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Carga Masiva.xlsx'
        archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

        logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
        logo.drawing.top = 10
        logo.drawing.left = 25

        hoja1 = archivo.get_sheet_by_name('hoja1')
        hoja1.title = "Carga Masiva Docentes"
        hoja1.add_image(logo)

        celda = hoja1.cell('E2')
        celda.value = 'Formacion'

        celda = hoja1.cell('E3')
        celda.value = 'Carga Masiva Docentes'

        celda = hoja1.cell('I3')
        celda.value = time.strftime("%d/%m/%y")

        celda = hoja1.cell('I4')
        celda.value = time.strftime("%I:%M:%S %p")

        row_num = 5

        columns = [tuple(['CODIGO GRUPO',30]),
                   tuple(['RADICADO',30]),
                   tuple(['NOMBRES',30]),
                   tuple(['APELLIDOS',30]),
                   tuple(['CEDULA',60]),
                   tuple(['CORREO',30]),
                   tuple(['TELEFONO FIJO',60]),
                   tuple(['CELULAR',30]),
                   tuple(['AREA',30]),
                   tuple(['GRADO',30]),
                   tuple(['TIPO BENEFICIARIO',30]),
                   tuple(['GENERO',30]),
                   tuple(['NOMBRE PROYECTO',30]),
                   tuple(['DEFINICION DEL PROBLEMA',30]),
                   tuple(['AREA PROYECTO',30]),
                   tuple(['COMPETENCIA',30]),
                   tuple(['GRUPO POBLACIONAL',30]),
                   tuple(['RESULTADO',30]),
                   ]

        for col_num in xrange(len(columns)):
            c = hoja1.cell(row=row_num, column=col_num+1)
            c.value = columns[col_num][0]
            c.style = t
            hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]


        archivo_masivo = openpyxl.load_workbook(settings.MEDIA_ROOT+'/'+unicode(archivo_queryset.archivo))

        hoja1_masivo = archivo_masivo.get_sheet_by_name('Hoja1')

        i = 0

        for fila in hoja1_masivo.rows:
            i += 1
            if i > 1:
                proceso =""
                try:
                    cedula = fila[4].value.replace('.','').replace(',','')
                except:
                    cedula = fila[4].value
                nombres = fila[2].value
                apellidos = fila[3].value


                if cedula == "" or cedula == None:
                    proceso = "El campo de cedula esta vacio"
                else:
                    try:
                        long(cedula)
                    except ValueError:
                        proceso = "El numero de cedula es invalido"
                    else:
                        if ParticipanteDocente.objects.filter(cedula=cedula).count() != 0:
                            proceso = "La Cedula ya se encuentra registrada"
                        else:
                            if nombres == "" or nombres == None:
                                proceso = "El campo de nombres esta vacio"
                            else:
                                if apellidos == "" or apellidos == None:
                                    proceso = "El campo de apellidos esta vacio"
                                else:
                                    if fila[0].value == "" or fila[3].value == None:
                                        proceso = "El Codigo de grupo no es valido"
                                    else:
                                        if GrupoDocentes.objects.filter(nombre=fila[0].value).count() != 1:
                                            proceso = "No hay un solo grupo registrado con el codigo de grupo"
                                        else:
                                            if RadicadoFormacion.objects.filter(numero=fila[1].value).count() != 1:
                                                proceso = "No hay un solo radicado con este numero"
                                            else:
                                                grupo = GrupoDocentes.objects.get(nombre=fila[0].value)
                                                proceso = "Registrado Correctamente"
                                                nuevo = ParticipanteDocente()
                                                nuevo.grupo = grupo
                                                nuevo.formador = grupo.formador
                                                nuevo.radicado = RadicadoFormacion.objects.get(numero=fila[1].value)

                                                nuevo.nombres = nombres
                                                nuevo.apellidos = apellidos
                                                nuevo.cedula = cedula
                                                if validateEmail(fila[5].value):
                                                    nuevo.correo = fila[5].value
                                                nuevo.telefono_fijo = fila[6].value
                                                nuevo.celular = fila[7].value

                                                if AreaCurricular.objects.filter(pk = fila[8].value).count()==1:
                                                    nuevo.area = AreaCurricular.objects.get(pk = fila[8].value)

                                                if Grado.objects.filter(pk = fila[9].value).count()==1:
                                                    nuevo.grado = Grado.objects.get(pk = fila[9].value)

                                                nuevo.tipo_beneficiario = fila[10].value

                                                if Genero.objects.filter(pk = fila[11].value).count()==1:
                                                    nuevo.genero = Genero.objects.get(pk = fila[11].value)

                                                nuevo.nombre_proyecto = fila[12].value
                                                nuevo.definicion_problema = fila[13].value
                                                nuevo.area_proyecto = fila[14].value

                                                if Competencias.objects.filter(pk = fila[15].value).count()==1:
                                                    nuevo.competencia = Competencias.objects.get(pk = fila[15].value)

                                                if GrupoPoblacional.objects.filter(pk = fila[16].value).count()==1:
                                                    nuevo.grupo_poblacional = GrupoPoblacional.objects.get(pk = fila[16].value)

                                                nuevo.save()

                row_num += 1
                row = [
                    fila[0].value,
                    fila[1].value,
                    fila[2].value,
                    fila[3].value,
                    fila[4].value,
                    fila[5].value,
                    fila[6].value,
                    fila[7].value,
                    fila[8].value,
                    fila[9].value,
                    fila[10].value,
                    fila[11].value,
                    fila[12].value,
                    fila[13].value,
                    fila[14].value,
                    fila[15].value,
                    fila[16].value,
                    proceso
                ]

                for col_num in xrange(len(row)):
                    c = hoja1.cell(row=row_num, column=col_num+1)
                    if row[col_num] == True:
                        c.value = "SI"
                    if row[col_num] == False:
                        c.value = "NO"
                    if row[col_num] == None:
                        c.value = ""
                    else:
                        c.value = row[col_num]
                    c.style = co

        archivo.save(response)
        return response
carga_docentes.short_description = "Cargar Docentes"

def actualizar_docentes(modeladmin,request,queryset):
    for archivo_queryset in queryset:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Actualizacion Docentes.xlsx'
        archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

        logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
        logo.drawing.top = 10
        logo.drawing.left = 25

        hoja1 = archivo.get_sheet_by_name('hoja1')
        hoja1.title = "Actualizacion Docentes"
        hoja1.add_image(logo)

        celda = hoja1.cell('E2')
        celda.value = 'Formacion'

        celda = hoja1.cell('E3')
        celda.value = 'Actualizacion Docentes'

        celda = hoja1.cell('I3')
        celda.value = time.strftime("%d/%m/%y")

        celda = hoja1.cell('I4')
        celda.value = time.strftime("%I:%M:%S %p")

        row_num = 5

        columns = [tuple(['CODIGO GRUPO',30]),
                   tuple(['RADICADO',30]),
                   tuple(['NOMBRES',30]),
                   tuple(['APELLIDOS',30]),
                   tuple(['CEDULA',60]),
                   tuple(['CORREO',30]),
                   tuple(['TELEFONO FIJO',60]),
                   tuple(['CELULAR',30]),
                   tuple(['AREA',30]),
                   tuple(['GRADO',30]),
                   tuple(['TIPO BENEFICIARIO',30]),
                   tuple(['GENERO',30]),
                   tuple(['NOMBRE PROYECTO',30]),
                   tuple(['DEFINICION DEL PROBLEMA',30]),
                   tuple(['AREA PROYECTO',30]),
                   tuple(['COMPETENCIA',30]),
                   tuple(['GRUPO POBLACIONAL',30]),
                   tuple(['RESULTADO',30]),
                   ]

        for col_num in xrange(len(columns)):
            c = hoja1.cell(row=row_num, column=col_num+1)
            c.value = columns[col_num][0]
            c.style = t
            hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]


        archivo_masivo = openpyxl.load_workbook(settings.MEDIA_ROOT+'/'+unicode(archivo_queryset.archivo))

        hoja1_masivo = archivo_masivo.get_sheet_by_name('Hoja1')

        i = 0

        for fila in hoja1_masivo.rows:
            i += 1
            if i > 1:
                proceso =""
                try:
                    cedula = fila[4].value.replace('.','').replace(',','')
                except:
                    cedula = fila[4].value
                nombres = fila[2].value
                apellidos = fila[3].value


                if cedula == "" or cedula == None:
                    proceso = "El campo de cedula esta vacio"
                else:
                    try:
                        long(cedula)
                    except ValueError:
                        proceso = "El numero de cedula es invalido"
                    else:
                        if ParticipanteDocente.objects.filter(cedula=cedula).count() == 1:
                            if nombres == "" or nombres == None:
                                proceso = "El campo de nombres esta vacio"
                            else:
                                if apellidos == "" or apellidos == None:
                                    proceso = "El campo de apellidos esta vacio"
                                else:
                                    nuevo = ParticipanteDocente.objects.get(cedula=cedula)

                                    if RadicadoFormacion.objects.filter(numero=fila[1].value).count() == 1:
                                        nuevo.radicado = RadicadoFormacion.objects.get(numero=fila[1].value)
                                    else:
                                        nuevo.radicado = RadicadoFormacion.objects.get(numero=0)

                                    nuevo.nombres = nombres
                                    nuevo.apellidos = apellidos

                                    if validateEmail(fila[5].value):
                                        nuevo.correo = fila[5].value

                                    if fila[6].value == "" or fila[6].value == None:
                                        nuevo.telefono_fijo = ""
                                    else:
                                        nuevo.telefono_fijo = fila[6].value

                                    if fila[7].value == "" or fila[7].value == None:
                                        nuevo.celular = ""
                                    else:
                                        nuevo.celular = fila[7].value

                                    if AreaCurricular.objects.filter(pk = fila[8].value).count()==1:
                                        nuevo.area = AreaCurricular.objects.get(pk = fila[8].value)

                                    if Grado.objects.filter(pk = fila[9].value).count()==1:
                                        nuevo.grado = Grado.objects.get(pk = fila[9].value)

                                    if fila[10].value == "" or fila[10].value == None:
                                        nuevo.tipo_beneficiario = ""
                                    else:
                                        nuevo.tipo_beneficiario = fila[10].value

                                    if Genero.objects.filter(genero = fila[11].value).count()==1:
                                        nuevo.genero = Genero.objects.get(genero = fila[11].value)

                                    if fila[12].value == "" or fila[12].value == None:
                                        nuevo.nombre_proyecto = ""
                                    else:
                                        nuevo.nombre_proyecto = fila[12].value

                                    if fila[13].value == "" or fila[13].value == None:
                                        nuevo.definicion_problema = ""
                                    else:
                                        nuevo.definicion_problema = fila[13].value

                                    if fila[14].value == "" or fila[14].value == None:
                                        nuevo.area_proyecto = ""
                                    else:
                                        nuevo.area_proyecto = fila[14].value

                                    if Competencias.objects.filter(pk = fila[15].value).count()==1:
                                        nuevo.competencia = Competencias.objects.get(pk = fila[15].value)

                                    if GrupoPoblacional.objects.filter(pk = fila[16].value).count()==1:
                                        nuevo.grupo_poblacional = GrupoPoblacional.objects.get(pk = fila[16].value)

                                    nuevo.save()


                row_num += 1
                row = [
                    fila[0].value,
                    fila[1].value,
                    fila[2].value,
                    fila[3].value,
                    fila[4].value,
                    fila[5].value,
                    fila[6].value,
                    fila[7].value,
                    fila[8].value,
                    fila[9].value,
                    fila[10].value,
                    fila[11].value,
                    fila[12].value,
                    fila[13].value,
                    fila[14].value,
                    fila[15].value,
                    fila[16].value,
                    proceso
                ]

                for col_num in xrange(len(row)):
                    c = hoja1.cell(row=row_num, column=col_num+1)
                    if row[col_num] == True:
                        c.value = "SI"
                    if row[col_num] == False:
                        c.value = "NO"
                    if row[col_num] == None:
                        c.value = ""
                    else:
                        c.value = row[col_num]
                    c.style = co

        archivo.save(response)
        return response
actualizar_docentes.short_description = "Actualizar Docentes"

def carga_participantes_truchos(modeladmin,request,queryset):
    for archivo_queryset in queryset:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Carga Masiva.xlsx'
        archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

        logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
        logo.drawing.top = 10
        logo.drawing.left = 25

        hoja1 = archivo.get_sheet_by_name('hoja1')
        hoja1.title = "Reporte Carga Masiva"
        hoja1.add_image(logo)

        celda = hoja1.cell('E2')
        celda.value = 'FORMACION'

        celda = hoja1.cell('E3')
        celda.value = 'Reporte Carga Masiva'

        celda = hoja1.cell('I3')
        celda.value = time.strftime("%d/%m/%y")

        celda = hoja1.cell('I4')
        celda.value = time.strftime("%I:%M:%S %p")

        row_num = 5

        columns = [tuple(['NOMBRE ARCHIVO',30]),
                   tuple(['INFORMACION',60]),
                   ]

        for col_num in xrange(len(columns)):
            c = hoja1.cell(row=row_num, column=col_num+1)
            c.value = columns[col_num][0]
            c.style = t
            hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]


        soportes = ZipFile(settings.MEDIA_ROOT+'//'+str(archivo_queryset.archivo),'r')

        name_list = soportes.namelist()

        for name in name_list:
            resultado = ""
            sesion = name.split("_")[0]
            cedula = name.split("_")[1].split(".")[0]

            try:
                participante_trucho = ParticipanteEscuelaTicTrucho.objects.get(cedula=cedula)
            except:
                resultado = "No existe el numero de cedula"
            else:
                cedulas = ParticipanteEscuelaTicTrucho.objects.filter(codigo_masivo__id=participante_trucho.codigo_masivo.id).values_list("cedula",flat=True)
                if sesion == "1":
                    resultado = "Participantes Cargados Exitosamente"
                    nuevo_soporte = SoporteEntregableEscuelaTic()
                    nuevo_soporte.grupo = participante_trucho.grupo
                    nuevo_soporte.entregable = Entregable.objects.get(id=5)
                    source = soportes.open(name)
                    target = file(os.path.join(r"C:\Temp",name),"wb")
                    with source, target:
                        shutil.copyfileobj(source,target)
                    nuevo_soporte.soporte = File(open("C://Temp//" + name, 'rb'))
                    nuevo_soporte.save()
                    os.remove("C://Temp//" + name)
                    for cedula in cedulas:
                        evidencia = EvidenciaEscuelaTic.objects.filter(participante__cedula=cedula).get(entregable=nuevo_soporte.entregable.id)
                        evidencia.soporte = nuevo_soporte

                        evidencia.save()
                elif sesion == "2":
                    resultado = "Participantes Cargados Exitosamente"
                    nuevo_soporte = SoporteEntregableEscuelaTic()
                    nuevo_soporte.grupo = participante_trucho.grupo
                    nuevo_soporte.entregable = Entregable.objects.get(id=9)
                    source = soportes.open(name)
                    target = file(os.path.join(r"C:\Temp",name),"wb")
                    with source, target:
                        shutil.copyfileobj(source,target)
                    nuevo_soporte.soporte = File(open("C://Temp//" + name, 'rb'))
                    nuevo_soporte.save()
                    os.remove("C://Temp//" + name)
                    for cedula in cedulas:
                        evidencia = EvidenciaEscuelaTic.objects.filter(participante__cedula=cedula).get(entregable=nuevo_soporte.entregable.id)
                        evidencia.soporte = nuevo_soporte
                        evidencia.save()
                else:
                    "El numero de sesion es invalido"


            row_num += 1
            row = [
                name,
                resultado
            ]

            for col_num in xrange(len(row)):
                c = hoja1.cell(row=row_num, column=col_num+1)
                if row[col_num] == True:
                    c.value = "SI"
                if row[col_num] == False:
                    c.value = "NO"
                if row[col_num] == None:
                    c.value = ""
                else:
                    c.value = row[col_num]
                c.style = co

        archivo.save(response)
        return response
carga_participantes_truchos.short_description = "Cargar Listados Masivos"


def actualizar_participantes(modeladmin,request,queryset):
    for archivo_queryset in queryset:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Actualizacion Masiva.xlsx'
        archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

        logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
        logo.drawing.top = 10
        logo.drawing.left = 25

        hoja1 = archivo.get_sheet_by_name('hoja1')
        hoja1.title = "Reporte Actualizacion Masiva"
        hoja1.add_image(logo)

        celda = hoja1.cell('E2')
        celda.value = 'FORMACION'

        celda = hoja1.cell('E3')
        celda.value = 'Reporte Actualizacion Masiva'

        celda = hoja1.cell('I3')
        celda.value = time.strftime("%d/%m/%y")

        celda = hoja1.cell('I4')
        celda.value = time.strftime("%I:%M:%S %p")

        row_num = 5

        columns = [tuple(['NOMBRE ARCHIVO',30]),
                   tuple(['INFORMACION',60]),
                   ]

        for col_num in xrange(len(columns)):
            c = hoja1.cell(row=row_num, column=col_num+1)
            c.value = columns[col_num][0]
            c.style = t
            hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]


        soportes = ZipFile(settings.MEDIA_ROOT+'//'+str(archivo_queryset.archivo),'r')

        name_list = soportes.namelist()

        for name in name_list:
            resultado = ""
            sesion = name.split("_")[1].split(".")[0]
            cedula = name.split("_")[0]

            try:
                participante_trucho = EvidenciaEscuelaTic.objects.filter(entregable__id=9).get(participante__cedula=cedula)
            except:
                resultado = "No existe el numero de cedula"
            else:
                if sesion == "2":
                    resultado = "Participantes Cargados Exitosamente"
                    nuevo_soporte = participante_trucho.soporte

                    source = soportes.open(name)
                    target = file(os.path.join(r"C:\Temp",name),"wb")
                    with source, target:
                        shutil.copyfileobj(source,target)

                    try:
                        nuevo_soporte.soporte = File(open("C://Temp//" + name, 'rb'))
                    except:
                        resultado = "Sin carga inicial"
                    else:
                        nuevo_soporte.save()
                        os.remove("C://Temp//" + name)
                else:
                    "El numero de sesion es invalido"


            row_num += 1
            row = [
                name,
                resultado
            ]

            for col_num in xrange(len(row)):
                c = hoja1.cell(row=row_num, column=col_num+1)
                if row[col_num] == True:
                    c.value = "SI"
                if row[col_num] == False:
                    c.value = "NO"
                if row[col_num] == None:
                    c.value = ""
                else:
                    c.value = row[col_num]
                c.style = co

        archivo.save(response)
        return response
actualizar_participantes.short_description = "Actualizar Listados Masivos"



def masivo_virtual(modeladmin,request,queryset):
    for archivo_queryset in queryset:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Carga Masiva.xlsx'
        archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

        logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
        logo.drawing.top = 10
        logo.drawing.left = 25

        hoja1 = archivo.get_sheet_by_name('hoja1')
        hoja1.title = "Reporte Carga Masiva"
        hoja1.add_image(logo)

        celda = hoja1.cell('E2')
        celda.value = 'FORMACION'

        celda = hoja1.cell('E3')
        celda.value = 'Reporte Carga Masiva'

        celda = hoja1.cell('I3')
        celda.value = time.strftime("%d/%m/%y")

        celda = hoja1.cell('I4')
        celda.value = time.strftime("%I:%M:%S %p")

        row_num = 5

        columns = [tuple(['CEDULA',30]),
                   tuple(['ID ACTIVIDAD',30]),
                   tuple(['PATH RELATIVO',30]),
                   tuple(['INFORMACION',60]),
                   ]

        for col_num in xrange(len(columns)):
            c = hoja1.cell(row=row_num, column=col_num+1)
            c.value = columns[col_num][0]
            c.style = t
            hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]


        soportes = ZipFile('C:\\Temp\\Carga.zip','r')

        archivo_masivo = openpyxl.load_workbook(settings.MEDIA_ROOT+'/'+unicode(archivo_queryset.archivo))

        hoja1_masivo = archivo_masivo.get_sheet_by_name('Hoja1')

        i = 0

        x = {'1':11,'2':11,'3':13,'4':15,'5':17,'6':17,'7':9,'8':9,'9':25,'10':26,'11':27,'12':35,'13':36,'14':37,
             '15':38,'16':39,'17':40,'18':51,'19':53,'20':55,'21':59,'22':60}

        for fila in hoja1_masivo.rows:
            i += 1
            if i > 1:
                proceso =""

                try:
                    evidencia_participante = EvidenciaDocentes.objects.filter(participante__cedula=fila[0].value)

                except:
                    proceso = "No existe el numero de cedula"

                else:
                    try:
                        evidencia = evidencia_participante.get(entregable__id=x[str(fila[1].value)])
                    except:
                        proceso = "No existe el espacio para la evidencia"
                    else:
                            path = fila[2].value
                            path = path.encode("CP850")
                            try:
                                info = soportes.getinfo(path)
                            except:
                                proceso = "No Existe el archivo en el path"
                            else:
                                proceso = "Soporte cargado"
                                filename = os.path.basename(path)
                                source = soportes.open(path)
                                target = file(os.path.join(r"C:\Temp",filename),"wb")
                                with source, target:
                                    shutil.copyfileobj(source,target)
                                nuevo_soporte = SoporteEntregableDocente()
                                nuevo_soporte.grupo = evidencia.participante.grupo
                                nuevo_soporte.entregable = evidencia.entregable

                                nuevo_soporte.soporte = File(open("C://Temp//" + filename, 'rb'))
                                nuevo_soporte.save()

                                evidencia.soporte = nuevo_soporte
                                evidencia.save()

                                #os.remove("C://Temp//" + filename)

                row_num += 1
                row = [
                    fila[0].value,
                    fila[1].value,
                    fila[2].value,
                    proceso
                ]

                for col_num in xrange(len(row)):
                    c = hoja1.cell(row=row_num, column=col_num+1)
                    if row[col_num] == True:
                        c.value = "SI"
                    if row[col_num] == False:
                        c.value = "NO"
                    if row[col_num] == None:
                        c.value = ""
                    else:
                        c.value = row[col_num]
                    c.style = co

        archivo.save(response)
        return response
masivo_virtual.short_description = "Carga masiva virtual"



def descarga_masiva(modeladmin,request,queryset):
    for archivo_queryset in queryset:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Descarga Masiva.xlsx'
        archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

        logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
        logo.drawing.top = 10
        logo.drawing.left = 25

        hoja1 = archivo.get_sheet_by_name('hoja1')
        hoja1.title = "Reporte Carga Masiva"
        hoja1.add_image(logo)

        celda = hoja1.cell('E2')
        celda.value = 'FORMACION'

        celda = hoja1.cell('E3')
        celda.value = 'Reporte Carga Masiva'

        celda = hoja1.cell('I3')
        celda.value = time.strftime("%d/%m/%y")

        celda = hoja1.cell('I4')
        celda.value = time.strftime("%I:%M:%S %p")

        row_num = 5

        columns = [tuple(['CEDULA',30]),
                   tuple(['ID ACTIVIDAD',30]),
                   tuple(['RESULTADO',30]),
                   ]

        for col_num in xrange(len(columns)):
            c = hoja1.cell(row=row_num, column=col_num+1)
            c.value = columns[col_num][0]
            c.style = t
            hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]


        archivo_masivo = openpyxl.load_workbook(settings.MEDIA_ROOT+'/'+unicode(archivo_queryset.archivo))

        hoja1_masivo = archivo_masivo.get_sheet_by_name('Hoja1')

        i = 0

        for fila in hoja1_masivo.rows:
            i += 1
            if i > 1:
                proceso = ""

                try:
                    evidencia = EvidenciaDocentes.objects.filter(participante__cedula=fila[0].value).get(entregable__id=fila[1].value)
                except:
                    proceso = "No existe el numero de cedula"
                else:
                    try:
                        path = evidencia.soporte.soporte.path
                    except:
                        proceso = "No hay evidencia cargada"
                    else:
                        proceso = path


                row_num += 1
                row = [
                    fila[0].value,
                    fila[1].value,
                    proceso
                ]

                for col_num in xrange(len(row)):
                    c = hoja1.cell(row=row_num, column=col_num+1)
                    if row[col_num] == True:
                        c.value = "SI"
                    if row[col_num] == False:
                        c.value = "NO"
                    if row[col_num] == None:
                        c.value = ""
                    else:
                        c.value = row[col_num]
                    c.style = co

        archivo.save(response)
        return response
descarga_masiva.short_description = "Descargar Masivamente"



class CargasMasivasAdmin(admin.ModelAdmin):
    list_display = ['id','archivo']
    ordering = ['archivo']
    actions = [carga_grupos,carga_participantes,carga_radicados,actualizar_radicados,carga_grupos_docentes,
               carga_docentes,actualizar_docentes,carga_participantes_truchos,actualizar_participantes,masivo_virtual,descarga_masiva]
admin.site.register(CargasMasivas, CargasMasivasAdmin)

def asignacion_total(modeladmin,request,queryset):
    for entregable in queryset:
        evidencias = EvidenciaDocentes.objects.all().filter(entregable__id=entregable.id).filter(soporte=None)
        for evidencia in evidencias:
            soporte = SoporteEntregableDocente.objects.filter(grupo__id=evidencia.participante.grupo.id).get(entregable__id=entregable.id)
            evidencia.soporte = soporte
            evidencia.save()
asignacion_total.short_description = "Asignar todos los participantes del grupo"

def crear_entregables_padres(modeladmin,request,queryset):
    participantes = ParticipanteEscuelaTic.objects.all().values_list('id',flat=True)
    for participante in participantes:
        if EvidenciaEscuelaTic.objects.filter(participante__id=participante).count() < 14:
            ids = EvidenciaEscuelaTic.objects.filter(participante__id=participante).values_list('entregable__id',flat=True)
            total_ids = Entregable.objects.all().values_list('id',flat=True)
            for id in total_ids:
                if id not in ids:
                    entregable = Entregable.objects.get(id=id)
                    p = ParticipanteEscuelaTic.objects.get(id=participante)
                    v = Valor.objects.get(id=1)

                    nuevo = EvidenciaEscuelaTic()
                    nuevo.soporte = None
                    nuevo.entregable = entregable
                    nuevo.participante = p
                    nuevo.valor = v
                    nuevo.corte = None
                    nuevo.usuario = None
                    nuevo.save()
crear_entregables_padres.short_description = "Verificar entregables padres"

def crear_entregables_docentes(modeladmin,request,queryset):
    participantes = ParticipanteDocente.objects.all().values_list('id',flat=True)
    for participante in participantes:
        if EvidenciaDocentes.objects.filter(participante__id=participante).count() < 33:
            ids = EvidenciaDocentes.objects.filter(participante__id=participante).values_list('entregable__id',flat=True)
            total_ids = EntregableDocentes.objects.all().values_list('id',flat=True)
            for id in total_ids:
                if id not in ids:
                    entregable = EntregableDocentes.objects.get(id=id)
                    p = ParticipanteDocente.objects.get(id=participante)
                    v = ValorDocente.objects.get(id=1)

                    nuevo = EvidenciaDocentes()
                    nuevo.soporte = None
                    nuevo.entregable = entregable
                    nuevo.participante = p
                    nuevo.valor = v
                    nuevo.corte = None
                    nuevo.usuario = None
                    nuevo.save()
crear_entregables_docentes.short_description = "Verificar entregables docentes"

class EntregableEscuelaTicAdmin(admin.ModelAdmin):
    list_display = ['nombre','actividad']
    ordering = ['id']
    actions = [asignacion_total,crear_entregables_padres]
admin.site.register(Entregable,EntregableEscuelaTicAdmin)

class EntregableDocentesAdmin(admin.ModelAdmin):
    list_display = ['nombre','actividad']
    ordering = ['id']
    actions = [asignacion_total,crear_entregables_docentes]
admin.site.register(EntregableDocentes,EntregableDocentesAdmin)



def reporte_docentes_revisados(modeladmin,request,queryset):
    for archivo_queryset in queryset:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Reporte Docentes.xlsx'
        archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

        logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
        logo.drawing.top = 10
        logo.drawing.left = 25

        hoja1 = archivo.get_sheet_by_name('hoja1')
        hoja1.title = "Reporte Docentes"
        hoja1.add_image(logo)

        celda = hoja1.cell('E2')
        celda.value = 'Formacion'

        celda = hoja1.cell('E3')
        celda.value = 'Reporte Docentes'

        celda = hoja1.cell('I3')
        celda.value = time.strftime("%d/%m/%y")

        celda = hoja1.cell('I4')
        celda.value = time.strftime("%I:%M:%S %p")

        row_num = 5

        columns = [tuple(['Region',30]),
                   tuple(['Nombres',30]),
                   tuple(['Apellidos',30]),
                   tuple(['Cedula',30]),
                   tuple(['Fecha',30]),
                   tuple(['Usuario',30]),
                   tuple(['Ip',30]),
                   ]

        for col_num in xrange(len(columns)):
            c = hoja1.cell(row=row_num, column=col_num+1)
            c.value = columns[col_num][0]
            c.style = t
            hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]



        for participante in RevisionInterventoriaDocente.objects.all():
                row_num += 1
                row = [
                    participante.region.nombre,
                    participante.participante.nombres,
                    participante.participante.apellidos,
                    participante.participante.cedula,
                    participante.fecha,
                    participante.usuario.username,
                    participante.ip,

                ]

                for col_num in xrange(len(row)):
                    c = hoja1.cell(row=row_num, column=col_num+1)
                    if row[col_num] == True:
                        c.value = "SI"
                    if row[col_num] == False:
                        c.value = "NO"
                    if row[col_num] == None:
                        c.value = ""
                    else:
                        c.value = row[col_num]
                    c.style = co

        archivo.save(response)
        return response
reporte_docentes_revisados.short_description = "Reporte Docentes"



def reporte_proyectos(modeladmin,request,queryset):
    for archivo_queryset in queryset:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Reporte Docentes.xlsx'
        archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

        logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
        logo.drawing.top = 10
        logo.drawing.left = 25

        hoja1 = archivo.get_sheet_by_name('hoja1')
        hoja1.title = "Reporte Proyectos"
        hoja1.add_image(logo)

        celda = hoja1.cell('E2')
        celda.value = 'Formacion'

        celda = hoja1.cell('E3')
        celda.value = 'Reporte Proyectos'

        celda = hoja1.cell('I3')
        celda.value = time.strftime("%d/%m/%y")

        celda = hoja1.cell('I4')
        celda.value = time.strftime("%I:%M:%S %p")

        row_num = 5

        columns = [tuple(['Region',30]),
                   tuple(['Formador',30]),
                   tuple(['Nombres',30]),
                   tuple(['Apellidos',30]),
                   tuple(['Cedula',30]),
                   ]

        for col_num in xrange(len(columns)):
            c = hoja1.cell(row=row_num, column=col_num+1)
            c.value = columns[col_num][0]
            c.style = t
            hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]



        for participante in EvidenciaDocentes.objects.filter(entregable__id=60).exclude(soporte__soporte=None):
                row_num += 1
                row = [
                    participante.participante.formador.region.nombre,
                    participante.participante.formador.nombre,
                    participante.participante.nombres,
                    participante.participante.apellidos,
                    participante.participante.cedula,
                ]

                for col_num in xrange(len(row)):
                    c = hoja1.cell(row=row_num, column=col_num+1)
                    if row[col_num] == True:
                        c.value = "SI"
                    if row[col_num] == False:
                        c.value = "NO"
                    if row[col_num] == None:
                        c.value = ""
                    else:
                        c.value = row[col_num]
                    c.style = co

        archivo.save(response)
        return response
reporte_proyectos.short_description = "Reporte Proyectos"



class RevisionInterventoriaDocenteAdmin(admin.ModelAdmin):
    list_display = ['participante']
    ordering = ['id']
    actions = [reporte_docentes_revisados,reporte_proyectos]

admin.site.register(RevisionInterventoriaDocente,RevisionInterventoriaDocenteAdmin)


def reporte_escuela_tic_revisados(modeladmin,request,queryset):
    for archivo_queryset in queryset:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Reporte Escuela Tic.xlsx'
        archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

        logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
        logo.drawing.top = 10
        logo.drawing.left = 25

        hoja1 = archivo.get_sheet_by_name('hoja1')
        hoja1.title = "Reporte Escuela Tic"
        hoja1.add_image(logo)

        celda = hoja1.cell('E2')
        celda.value = 'Formacion'

        celda = hoja1.cell('E3')
        celda.value = 'Reporte Escuela Tic'

        celda = hoja1.cell('I3')
        celda.value = time.strftime("%d/%m/%y")

        celda = hoja1.cell('I4')
        celda.value = time.strftime("%I:%M:%S %p")

        row_num = 5

        columns = [tuple(['Region',30]),
                   tuple(['Nombres',30]),
                   tuple(['Apellidos',30]),
                   tuple(['Cedula',30]),
                   tuple(['Fecha',30]),
                   tuple(['Usuario',30]),
                   tuple(['Ip',30]),
                   ]

        for col_num in xrange(len(columns)):
            c = hoja1.cell(row=row_num, column=col_num+1)
            c.value = columns[col_num][0]
            c.style = t
            hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]



        for participante in RevisionInterventoriaEscuelaTic.objects.all():
                row_num += 1
                row = [
                    participante.region.nombre,
                    participante.participante.nombres,
                    participante.participante.apellidos,
                    participante.participante.cedula,
                    participante.fecha,
                    participante.usuario.username,
                    participante.ip,
                ]

                for col_num in xrange(len(row)):
                    c = hoja1.cell(row=row_num, column=col_num+1)
                    if row[col_num] == True:
                        c.value = "SI"
                    if row[col_num] == False:
                        c.value = "NO"
                    if row[col_num] == None:
                        c.value = ""
                    else:
                        c.value = row[col_num]
                    c.style = co

        archivo.save(response)
        return response
reporte_escuela_tic_revisados.short_description = "Reporte Escuela Tic"

def matriz_escuela_tic(modeladmin,request,queryset):

    data = ParticipanteEscuelaTic.objects.all()
    chunks=[data[x:x+35000] for x in xrange(0, data.count(), 35000)]
    i = 0
    for chunk in chunks:
        i += 1
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Matriz Escuela Tic '+unicode(i)+'.xlsx'
        archivo = openpyxl.load_workbook(settings.STATICFILES_DIRS[0]+'/formatos/base.xlsx')

        logo = openpyxl.drawing.Image(settings.STATICFILES_DIRS[0]+'/formatos/logo.png')
        logo.drawing.top = 10
        logo.drawing.left = 25

        hoja1 = archivo.get_sheet_by_name('hoja1')
        hoja1.title = "Matriz Escuela Tic"
        hoja1.add_image(logo)

        celda = hoja1.cell('E2')
        celda.value = 'Formacion'

        celda = hoja1.cell('E3')
        celda.value = 'Matriz Escuela Tic'

        celda = hoja1.cell('I3')
        celda.value = time.strftime("%d/%m/%y")

        celda = hoja1.cell('I4')
        celda.value = time.strftime("%I:%M:%S %p")

        row_num = 5

        columns = [tuple(['Region',30]),
                   tuple(['Departamento',30]),
                   tuple(['Municipio',30]),
                   tuple(['Institucion',30]),
                   tuple(['Codigo del Grupo',30]),
                   tuple(['Nombre del Formador',30]),
                   tuple(['Cedula del Formador',30]),
                   tuple(['Nombres',30]),
                   tuple(['Apellidos',30]),
                   tuple(['Cedula',30]),
                   tuple(['Genero',30]),
                   tuple(['Nivel Educativo',30]),
                   tuple(['Telefono',30]),
                   tuple(['Correo',30]),
                   tuple(['Tipo Poblacion',30]),
                   tuple(['Codigo Proyecto',30]),
                   tuple(['Tipo Proyecto',30]),
                   tuple(['Grupo Conformacion',30]),
                   tuple(['Sesion 1',30]),
                   tuple(['Sesion 2',30]),
                   ]

        for col_num in xrange(len(columns)):
            c = hoja1.cell(row=row_num, column=col_num+1)
            c.value = columns[col_num][0]
            c.style = t
            hoja1.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = columns[col_num][1]


        for participante in chunk:
                row_num += 1
                try:
                    sesion_1 = unicode(EvidenciaEscuelaTic.objects.filter(participante__id=participante.id).get(entregable__id=5).soporte.soporte)
                except:
                    sesion_1 = ""
                try:
                    sesion_2 = unicode(EvidenciaEscuelaTic.objects.filter(participante__id=participante.id).get(entregable__id=9).soporte.soporte)
                except:
                    sesion_2 = ""
                row = [
                    participante.formador.region.nombre,
                    participante.grupo.municipio.departamento.nombre,
                    participante.grupo.municipio.nombre,
                    participante.institucion,
                    participante.grupo.nombre,
                    participante.formador.nombre,
                    participante.formador.cedula,
                    participante.nombres,
                    participante.apellidos,
                    participante.cedula,
                    participante.genero,
                    participante.nivel_educativo,
                    participante.telefono,
                    participante.correo,
                    participante.poblacion,
                    participante.codigo_anspe,
                    participante.tipo_proyecto,
                    participante.grupo_conformacion,
                    sesion_1,
                    sesion_2
                ]

                for col_num in xrange(len(row)):
                    c = hoja1.cell(row=row_num, column=col_num+1)
                    if row[col_num] == True:
                        c.value = "SI"
                    if row[col_num] == False:
                        c.value = "NO"
                    if row[col_num] == None:
                        c.value = ""
                    else:
                        c.value = row[col_num]
                    c.style = co

        archivo.save(settings.MEDIA_ROOT+'/Matriz Padres/Matriz'+unicode(i)+'.xlsx')
    return HttpResponseRedirect('/media/Matriz Padres/Matriz.xlsx')
matriz_escuela_tic.short_description = "Matriz Escuela Tic"

class RevisionInterventoriaEscuelaTicAdmin(admin.ModelAdmin):
    list_display = ['participante']
    ordering = ['id']
    actions = [reporte_escuela_tic_revisados,matriz_escuela_tic]

admin.site.register(RevisionInterventoriaEscuelaTic,RevisionInterventoriaEscuelaTicAdmin)
admin.site.register(RevisionInterventoriaDocenteSoporte)
admin.site.register(RevisionInterventoriaEscuelaTicSoporte)
admin.site.register(RevisionInterventoriaEscuelaTicSoporteActividades)
admin.site.register(RevisionInterventoriaDocenteSoporteActividades)
admin.site.register(AreaCurricular)
admin.site.register(Grado)
admin.site.register(Genero)
admin.site.register(Competencias)
admin.site.register(GrupoPoblacional)




class RevisionInterventoriaEscuelaTicAdmin(admin.ModelAdmin):
    list_display = ['participante']
    ordering = ['id']
    actions = [reporte_escuela_tic_revisados,matriz_escuela_tic]
admin.site.register(CargaMasivaListados)